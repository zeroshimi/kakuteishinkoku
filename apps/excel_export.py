"""Excel出力（副業種別ごとにシート分け）"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from models import Transaction
from config import ACCOUNTS, get_side_job_types


def _sheet_name(name: str) -> str:
    return name[:31] if len(name) > 31 else name


def _entry_to_row(t: Transaction) -> tuple:
    lines = {l.side: l for l in t.lines}
    debit = lines.get("debit")
    credit = lines.get("credit")
    return (
        t.date.strftime("%Y-%m-%d"),
        debit.account if debit else "",
        credit.account if credit else "",
        debit.amount if debit else 0,
        credit.amount if credit else 0,
        t.description or "",
        "○" if t.receipt_file else "",
    )


def _write_entries_sheet(wb: Workbook, sheet_title: str, entries: list, side_job_label: str = "") -> None:
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    headers = ["日付", "借方科目", "貸方科目", "借方金額", "貸方金額", "摘要", "領収書"]

    ws = wb.create_sheet(_sheet_name(sheet_title))
    row = 1
    if side_job_label:
        ws.cell(row=row, column=1, value=f"副業種別: {side_job_label}").font = Font(bold=True, size=12)
        row += 2

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    for t in entries:
        r = _entry_to_row(t)
        for col, val in enumerate(r, 1):
            ws.cell(row=row, column=col, value=val).border = thin_border
        row += 1

    for i in range(1, 8):
        w = 14 if i <= 4 else (30 if i == 6 else 8)
        ws.column_dimensions[get_column_letter(i)].width = w


def _calc_zashotoku(entries: list) -> float:
    """雑所得 = 収入 - 必要経費"""
    revenue_accounts = [a[0] for a in ACCOUNTS if a[1] == "revenue"]
    expense_accounts = [a[0] for a in ACCOUNTS if a[1] == "expense"]
    total_revenue = 0.0
    total_expense = 0.0
    for t in entries:
        for l in t.lines:
            if l.side == "credit" and l.account in revenue_accounts:
                total_revenue += l.amount
            if l.side == "debit" and l.account in expense_accounts:
                total_expense += l.amount
    return total_revenue - total_expense


def _calc_totals_by_account(entries: list) -> dict:
    """科目別の収入・経費を集計"""
    revenue_accounts = [a[0] for a in ACCOUNTS if a[1] == "revenue"]
    expense_accounts = [a[0] for a in ACCOUNTS if a[1] == "expense"]
    totals = {}
    for t in entries:
        for l in t.lines:
            if l.side == "credit" and l.account in revenue_accounts:
                totals[l.account] = totals.get(l.account, 0) + l.amount
            if l.side == "debit" and l.account in expense_accounts:
                totals[l.account] = totals.get(l.account, 0) + l.amount
    return totals


def _write_zashotoku_summary_sheet(wb: Workbook, year: int, all_entries: list) -> None:
    """雑所得の合計と副業ごとの集計（経費も科目ごと）"""
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    expense_accounts = [a[0] for a in ACCOUNTS if a[1] == "expense"]
    revenue_accounts = [a[0] for a in ACCOUNTS if a[1] == "revenue"]

    ws = wb.create_sheet(_sheet_name(f"{year}年度 雑所得サマリー"))
    ws.cell(row=1, column=1, value="雑所得の集計（副業ごと・科目ごと）").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"{year}年度").font = Font(bold=True)
    row = 4
    total_zashotoku = 0.0
    for sj in get_side_job_types():
        entries = [e for e in all_entries if e.side_job_type == sj]
        totals = _calc_totals_by_account(entries)
        total_revenue = sum(totals.get(a, 0) for a in revenue_accounts)
        total_expense = sum(totals.get(a, 0) for a in expense_accounts)
        z = total_revenue - total_expense
        total_zashotoku += z

        ws.cell(row=row, column=1, value=f"【{sj}】").font = header_font
        row += 1
        ws.cell(row=row, column=1, value="科目").font = header_font
        ws.cell(row=row, column=2, value="金額").font = header_font
        row += 1
        for acc in expense_accounts:
            val = totals.get(acc, 0)
            if val != 0:
                ws.cell(row=row, column=1, value=acc).border = thin_border
                ws.cell(row=row, column=2, value=val).border = thin_border
                row += 1
        for acc in revenue_accounts:
            val = totals.get(acc, 0)
            if val != 0:
                ws.cell(row=row, column=1, value=acc).border = thin_border
                ws.cell(row=row, column=2, value=val).border = thin_border
                row += 1
        ws.cell(row=row, column=1, value="必要経費の計").font = header_font
        ws.cell(row=row, column=2, value=total_expense).font = header_font
        row += 1
        ws.cell(row=row, column=1, value="雑所得").font = header_font
        ws.cell(row=row, column=2, value=z).font = header_font
        row += 2

    ws.cell(row=row, column=1, value="合計 雑所得").font = Font(bold=True, size=12)
    ws.cell(row=row, column=2, value=total_zashotoku).font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14


# 科目の出力順（ユーザー指定）
SUMMARY_ACCOUNT_ORDER = [
    "立替金", "事業主貸", "売掛金", "現金",
    "接待交際費", "会議費", "消耗品費", "旅費交通費", "地代家賃", "水道光熱費",
    "通信費", "支払手数料", "新聞図書費", "研修費", "雑費", "租税公課", "減価償却費",
    "事業主借", "売上高",
]


def _write_summary_sheet(wb: Workbook, year: int, entries: list, title_suffix: str = "全体") -> None:
    """科目別の収入・経費・雑所得を副業ごとに出力（指定順）"""
    debit_accounts = [a[0] for a in ACCOUNTS if a[1] in ("asset", "expense")]
    credit_accounts = [a[0] for a in ACCOUNTS if a[1] in ("liability", "revenue")]
    expense_accounts = [a[0] for a in ACCOUNTS if a[1] == "expense"]
    revenue_accounts = [a[0] for a in ACCOUNTS if a[1] == "revenue"]
    totals = {name: 0.0 for name in SUMMARY_ACCOUNT_ORDER}

    for t in entries:
        for l in t.lines:
            if l.account not in totals:
                continue
            if l.side == "debit" and l.account in debit_accounts:
                totals[l.account] += l.amount
            elif l.side == "credit" and l.account in credit_accounts:
                totals[l.account] += l.amount
            elif l.account == "売掛金":
                if l.side == "debit":
                    totals[l.account] += l.amount
                else:
                    totals[l.account] -= l.amount

    header_font = Font(bold=True)
    sheet_title = f"{year}年度 集計_{title_suffix}" if title_suffix != "全体" else f"{year}年度 集計"
    ws = wb.create_sheet(_sheet_name(sheet_title))
    ws.cell(row=1, column=1, value="副業に係る雑所得の金額の計算表").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"{year}年度（{title_suffix}）").font = Font(bold=True)
    ws.cell(row=4, column=1, value="科目").font = header_font
    ws.cell(row=4, column=2, value="金額").font = header_font
    row_num = 5
    total_expense = sum(totals.get(a, 0) for a in expense_accounts)
    total_revenue = sum(totals.get(a, 0) for a in revenue_accounts)
    for name in SUMMARY_ACCOUNT_ORDER:
        val = totals.get(name, 0)
        ws.cell(row=row_num, column=1, value=name)
        ws.cell(row=row_num, column=2, value=val if val != 0 else "")
        row_num += 1
    ws.cell(row=row_num, column=1, value="必要経費の計").font = header_font
    ws.cell(row=row_num, column=2, value=total_expense).font = header_font
    row_num += 1
    ws.cell(row=row_num, column=1, value="雑所得の金額").font = header_font
    ws.cell(row=row_num, column=2, value=total_revenue - total_expense).font = header_font
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14


def export_to_excel(db: Session, year: int) -> str:
    all_entries = (
        db.query(Transaction)
        .options(joinedload(Transaction.lines))
        .filter(Transaction.year == year)
        .order_by(Transaction.date, Transaction.id)
        .all()
    )

    wb = Workbook()
    wb.remove(wb.active)
    _write_zashotoku_summary_sheet(wb, year, all_entries)
    _write_summary_sheet(wb, year, all_entries, "全体")
    for sj in get_side_job_types():
        entries = [e for e in all_entries if e.side_job_type == sj]
        _write_summary_sheet(wb, year, entries, sj)
    for sj in get_side_job_types():
        entries = [e for e in all_entries if e.side_job_type == sj]
        _write_entries_sheet(wb, f"{sj}_一覧", entries, sj)

    output_dir = Path(__file__).parent / "output"
    output_dir.mkdir(exist_ok=True)
    file_path = output_dir / f"複式簿記_{year}年度.xlsx"
    wb.save(file_path)
    return str(file_path)
